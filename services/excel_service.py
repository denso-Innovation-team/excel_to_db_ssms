"""
services/excel_service.py
Excel Processing Service - Clean & Focused - FIXED
"""

from typing import Dict, Any, List, Tuple
import pandas as pd
from pathlib import Path
from datetime import datetime
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

logger = logging.getLogger(__name__)


class ExcelService:
    """Excel file processing service"""

    def __init__(self):
        self.supported_extensions = [".xlsx", ".xls", ".xlsm", ".xlsb"]
        self.current_file = None
        self.file_info = {}

    def analyze_file(self, file_path: str) -> Dict[str, Any]:
        """Analyze Excel file and return basic information"""
        try:
            file_path = Path(file_path)

            # Validate file
            if not self._validate_file(file_path):
                return {"error": "Invalid file or unsupported format"}

            # Read basic info
            info = self._get_basic_info(file_path)

            # Read sample data
            sample_data = self._read_sample_data(file_path)
            info.update(sample_data)

            self.current_file = str(file_path)
            self.file_info = info

            return info

        except Exception as e:
            logger.error(f"Failed to analyze Excel file: {e}")
            return {"error": str(e)}

    def read_file(self, file_path: str, options: Dict[str, Any] = None) -> List[Dict]:
        """Read Excel file and return data as list of dictionaries"""
        try:
            options = options or {}

            # Read Excel file
            df = pd.read_excel(
                file_path,
                header=0 if options.get("has_header", True) else None,
                sheet_name=options.get("sheet_name", 0),
            )

            # Clean data if requested
            if options.get("clean_data", True):
                df = self._clean_dataframe(df)

            # Convert to list of dictionaries
            data = df.to_dict("records")

            logger.info(f"Successfully read {len(data)} rows from Excel file")
            return data

        except Exception as e:
            logger.error(f"Failed to read Excel file: {e}")
            raise Exception(f"Excel read error: {str(e)}")

    def export_data(
        self, data: List[Dict], file_path: str, format_type: str = "xlsx"
    ) -> bool:
        """Export data to Excel file"""
        try:
            if not data:
                raise ValueError("No data to export")

            df = pd.DataFrame(data)

            if format_type.lower() == "xlsx":
                df.to_excel(file_path, index=False, engine="openpyxl")
            elif format_type.lower() == "csv":
                df.to_csv(file_path, index=False, encoding="utf-8")
            else:
                raise ValueError(f"Unsupported export format: {format_type}")

            logger.info(f"Successfully exported {len(data)} rows to {file_path}")
            return True

        except Exception as e:
            logger.error(f"Failed to export data: {e}")
            return False

    def get_sheet_names(self, file_path: str) -> List[str]:
        """Get list of sheet names in Excel file"""
        try:
            excel_file = pd.ExcelFile(file_path)
            return excel_file.sheet_names
        except Exception as e:
            logger.error(f"Failed to get sheet names: {e}")
            return []

    def validate_file(self, file_path: str) -> Tuple[bool, str]:
        """Validate Excel file"""
        try:
            file_path = Path(file_path)

            if not file_path.exists():
                return False, "File does not exist"

            if file_path.suffix.lower() not in self.supported_extensions:
                return (
                    False,
                    f"Unsupported file type. Supported: {', '.join(self.supported_extensions)}",
                )

            # Try to read first row
            pd.read_excel(file_path, nrows=1)

            return True, "File is valid"

        except Exception as e:
            return False, f"File validation failed: {str(e)}"

    def get_file_info(self, file_path: str) -> Dict[str, Any]:
        """Get comprehensive file information"""
        try:
            file_path = Path(file_path)

            # Basic file info
            stat = file_path.stat()
            file_size_mb = round(stat.st_size / (1024 * 1024), 2)
            modified_date = datetime.fromtimestamp(stat.st_mtime)

            # Excel info
            df = pd.read_excel(file_path)

            return {
                "file_name": file_path.name,
                "file_path": str(file_path.absolute()),
                "file_size_mb": file_size_mb,
                "modified_date": modified_date.strftime("%Y-%m-%d %H:%M:%S"),
                "total_rows": len(df),
                "total_columns": len(df.columns),
                "columns": list(df.columns),
                "data_types": {col: str(dtype) for col, dtype in df.dtypes.items()},
                "sample_data": df.head(3).to_dict("records"),
                "sheet_names": self.get_sheet_names(file_path),
            }

        except Exception as e:
            logger.error(f"Failed to get file info: {e}")
            return {"error": str(e)}

    def _validate_file(self, file_path: Path) -> bool:
        """Internal file validation"""
        if not file_path.exists():
            return False

        if file_path.suffix.lower() not in self.supported_extensions:
            return False

        try:
            # Quick read test
            pd.read_excel(file_path, nrows=1)
            return True
        except Exception:
            return False

    def _get_basic_info(self, file_path: Path) -> Dict[str, Any]:
        """Get basic file information"""
        stat = file_path.stat()

        return {
            "file_name": file_path.name,
            "file_path": str(file_path.absolute()),
            "file_size_mb": round(stat.st_size / (1024 * 1024), 2),
            "modified_date": datetime.fromtimestamp(stat.st_mtime).strftime(
                "%Y-%m-%d %H:%M:%S"
            ),
        }

    def _read_sample_data(
        self, file_path: Path, sample_rows: int = 10
    ) -> Dict[str, Any]:
        """Read sample data from Excel file"""
        try:
            # Read sample rows
            df_sample = pd.read_excel(file_path, nrows=sample_rows)

            # Read full file for row count (more efficient way)
            df_info = pd.read_excel(file_path, usecols=[0])
            total_rows = len(df_info)

            return {
                "total_rows": total_rows,
                "total_columns": len(df_sample.columns),
                "columns": list(df_sample.columns),
                "data_types": {
                    col: str(dtype) for col, dtype in df_sample.dtypes.items()
                },
                "sample_data": df_sample.head(5).to_dict("records"),
                "data_quality": self._analyze_data_quality(df_sample),
            }

        except Exception as e:
            logger.error(f"Failed to read sample data: {e}")
            return {"error": str(e)}

    def _clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """Clean DataFrame for database import"""
        try:
            # Remove completely empty rows
            df = df.dropna(how="all")

            # Clean column names
            df.columns = [self._clean_column_name(col) for col in df.columns]

            # Handle missing values
            for col in df.columns:
                if df[col].dtype == "object":
                    df[col] = df[col].fillna("").astype(str).str.strip()
                else:
                    df[col] = df[col].fillna(0)

            # Remove duplicate rows
            df = df.drop_duplicates()

            return df

        except Exception as e:
            logger.error(f"Failed to clean DataFrame: {e}")
            return df

    def _clean_column_name(self, name: str) -> str:
        """Clean column name for database compatibility"""
        import re

        # Convert to string and strip
        clean_name = str(name).strip()

        # Replace special characters with underscore
        clean_name = re.sub(r"[^\w\s]", "_", clean_name)

        # Replace spaces with underscore
        clean_name = re.sub(r"\s+", "_", clean_name)

        # Remove multiple underscores
        clean_name = re.sub(r"_+", "_", clean_name)

        # Remove leading/trailing underscores and convert to lowercase
        clean_name = clean_name.strip("_").lower()

        # Ensure not empty
        if not clean_name:
            clean_name = "column"

        # Handle reserved keywords
        reserved_words = {"index", "order", "group", "select", "from", "where", "table"}
        if clean_name in reserved_words:
            clean_name = f"{clean_name}_col"

        return clean_name

    def _analyze_data_quality(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Analyze data quality metrics"""
        try:
            total_cells = len(df) * len(df.columns)
            null_cells = df.isnull().sum().sum()

            return {
                "total_rows": len(df),
                "total_columns": len(df.columns),
                "null_percentage": (
                    round((null_cells / total_cells) * 100, 2) if total_cells > 0 else 0
                ),
                "duplicate_rows": df.duplicated().sum(),
                "empty_columns": df.isnull().all().sum(),
                "mixed_type_columns": self._detect_mixed_types(df),
                "potential_issues": self._detect_potential_issues(df),
            }

        except Exception as e:
            logger.error(f"Failed to analyze data quality: {e}")
            return {"error": str(e)}

    def _detect_mixed_types(self, df: pd.DataFrame) -> List[str]:
        """Detect columns with mixed data types"""
        mixed_columns = []

        for col in df.columns:
            if df[col].dtype == "object":
                # Check for mixed numeric/text data
                non_null_data = df[col].dropna()
                if len(non_null_data) > 0:
                    numeric_count = 0
                    for value in non_null_data:
                        try:
                            float(str(value))
                            numeric_count += 1
                        except (ValueError, TypeError):
                            pass

                    # If partially numeric, it might be mixed type
                    if 0 < numeric_count < len(non_null_data):
                        mixed_columns.append(col)

        return mixed_columns

    def _detect_potential_issues(self, df: pd.DataFrame) -> List[str]:
        """Detect potential data issues"""
        issues = []

        # Check for very long strings that might cause database issues
        for col in df.select_dtypes(include=["object"]).columns:
            max_length = df[col].astype(str).str.len().max()
            if max_length > 1000:
                issues.append(
                    f"Column '{col}' has very long text (max: {max_length} chars)"
                )

        # Check for columns with mostly null values
        for col in df.columns:
            null_percentage = (df[col].isnull().sum() / len(df)) * 100
            if null_percentage > 90:
                issues.append(
                    f"Column '{col}' is mostly empty ({null_percentage:.1f}% null)"
                )

        # Check for potential date columns that aren't recognized
        for col in df.select_dtypes(include=["object"]).columns:
            sample_values = df[col].dropna().head(10)
            date_like_count = 0
            for value in sample_values:
                if self._looks_like_date(str(value)):
                    date_like_count += 1

            if date_like_count > len(sample_values) * 0.7:
                issues.append(
                    f"Column '{col}' might contain dates but wasn't recognized"
                )

        return issues

    def _looks_like_date(self, value: str) -> bool:
        """Check if string looks like a date"""
        import re

        date_patterns = [
            r"\d{4}-\d{2}-\d{2}",  # YYYY-MM-DD
            r"\d{2}/\d{2}/\d{4}",  # MM/DD/YYYY or DD/MM/YYYY
            r"\d{2}-\d{2}-\d{4}",  # MM-DD-YYYY or DD-MM-YYYY
            r"\d{1,2}/\d{1,2}/\d{4}",  # M/D/YYYY
        ]

        for pattern in date_patterns:
            if re.match(pattern, value.strip()):
                return True

        return False

    def create_template(
        self,
        template_type: str,
        output_path: str,
        rows: int = 100,
    ) -> Dict[str, Any]:  # Change return type to Dict
        """Create Excel template with improved functionality"""
        templates = {
            "employees": {
                "columns": [
                    "employee_id",
                    "first_name",
                    "last_name",
                    "email",
                    "department",
                    "position",
                    "salary",
                    "hire_date",
                    "active",
                ],
                "sample_data": [
                    [
                        "EMP001",
                        "John",
                        "Doe",
                        "john.doe@company.com",
                        "IT",
                        "Developer",
                        50000,
                        "2023-01-15",
                        True,
                    ],
                    [
                        "EMP002",
                        "Jane",
                        "Smith",
                        "jane.smith@company.com",
                        "HR",
                        "Manager",
                        65000,
                        "2022-06-10",
                        True,
                    ],
                    [
                        "EMP003",
                        "สมชาย",
                        "ใจดี",
                        "somchai@company.com",
                        "Production",
                        "Operator",
                        35000,
                        "2023-03-20",
                        True,
                    ],
                ],
                "validations": {
                    "employee_id": "Must be unique",
                    "email": "Valid email format required",
                    "salary": "Numeric value > 0",
                    "hire_date": "YYYY-MM-DD format",
                },
            },
            "sales": {
                "columns": [
                    "transaction_id",
                    "customer_name",
                    "product",
                    "quantity",
                    "unit_price",
                    "total_amount",
                    "sale_date",
                ],
                "sample_data": [
                    [
                        "TXN001",
                        "ABC Corp",
                        "Widget A",
                        10,
                        25.50,
                        255.00,
                        "2024-01-01",
                    ],
                    [
                        "TXN002",
                        "XYZ Ltd",
                        "Gadget B",
                        5,
                        100.00,
                        500.00,
                        "2024-01-02",
                    ],
                    [
                        "TXN003",
                        "DEF Inc",
                        "Tool C",
                        15,
                        15.75,
                        236.25,
                        "2024-01-03",
                    ],
                ],
            },
            "inventory": {
                "columns": [
                    "product_id",
                    "product_name",
                    "category",
                    "current_stock",
                    "unit_price",
                    "supplier",
                    "warehouse",
                ],
                "sample_data": [
                    [
                        "PROD001",
                        "Engine Part A1",
                        "Auto Parts",
                        150,
                        500.00,
                        "DENSO",
                        "Bangkok",
                    ],
                    [
                        "PROD002",
                        "Brake System B2",
                        "Brake Parts",
                        85,
                        1200.00,
                        "Bosch",
                        "Chonburi",
                    ],
                    [
                        "PROD003",
                        "ECU Module C3",
                        "Electronics",
                        45,
                        2500.00,
                        "Continental",
                        "Rayong",
                    ],
                ],
            },
        }

        try:
            template = templates.get(template_type)
            if not template:
                return {
                    "success": False,
                    "message": "Template type not found",
                }

            # Create workbook with validation
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"

            # Add headers
            for col, header in enumerate(template["columns"], 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E2E8F0", fill_type="solid")

            # Add format guide
            guide_sheet = wb.create_sheet("Guide")
            guide_sheet.append(["Column", "Format", "Validation"])
            for col in template["columns"]:
                guide_sheet.append(
                    [
                        col,
                        template.get("formats", {}).get(col, "Text"),
                        template["validations"].get(col, ""),
                    ]
                )

            # Save template
            wb.save(output_path)
            return {
                "success": True,
                "message": "Template created successfully",
                "file_path": output_path,
            }

        except Exception as e:
            return {
                "success": False,
                "message": f"Error creating template: {str(e)}",
            }

    def batch_process_files(
        self, file_paths: List[str], output_dir: str
    ) -> Dict[str, Any]:
        """Process multiple Excel files in batch"""
        results = {
            "processed": [],
            "failed": [],
            "total_files": len(file_paths),
            "total_rows": 0,
        }

        for file_path in file_paths:
            try:
                # Analyze file
                file_info = self.analyze_file(file_path)

                if "error" in file_info:
                    results["failed"].append(
                        {"file": file_path, "error": file_info["error"]}
                    )
                    continue

                # Read and process
                data = self.read_file(file_path)

                # Generate output filename
                input_path = Path(file_path)
                output_file = Path(output_dir) / f"processed_{input_path.stem}.xlsx"

                # Export processed data
                if self.export_data(data, str(output_file)):
                    results["processed"].append(
                        {
                            "input_file": file_path,
                            "output_file": str(output_file),
                            "rows": len(data),
                            "columns": len(data[0]) if data else 0,
                        }
                    )
                    results["total_rows"] += len(data)
                else:
                    results["failed"].append(
                        {"file": file_path, "error": "Export failed"}
                    )

            except Exception as e:
                results["failed"].append({"file": file_path, "error": str(e)})

        return results

    def get_column_suggestions(self, file_path: str) -> Dict[str, str]:
        """Get column type suggestions for database import"""
        try:
            df = pd.read_excel(file_path, nrows=100)  # Sample for analysis
            suggestions = {}

            for col in df.columns:
                clean_col = self._clean_column_name(col)

                # Analyze data type
                non_null_data = df[col].dropna()
                if len(non_null_data) == 0:
                    suggestions[clean_col] = "TEXT"
                    continue

                # Check for numeric data
                try:
                    pd.to_numeric(non_null_data)
                    # Check if integers
                    if non_null_data.astype(str).str.match(r"^\d+$").all():
                        suggestions[clean_col] = "INTEGER"
                    else:
                        suggestions[clean_col] = "REAL"
                    continue
                except Exception:
                    pass

                # Check for dates
                try:
                    pd.to_datetime(non_null_data)
                    suggestions[clean_col] = "DATE"
                    continue
                except Exception:
                    pass

                # Check for boolean
                unique_vals = set(str(v).lower() for v in non_null_data.unique())
                bool_vals = {"true", "false", "yes", "no", "1", "0", "y", "n"}
                if unique_vals.issubset(bool_vals):
                    suggestions[clean_col] = "BOOLEAN"
                    continue

                # Default to text
                suggestions[clean_col] = "TEXT"

            return suggestions

        except Exception as e:
            logger.error(f"Failed to get column suggestions: {e}")
            return {}


# Utility functions
def quick_excel_info(file_path: str) -> Dict[str, Any]:
    """Quick way to get Excel file info"""
    service = ExcelService()
    return service.get_file_info(file_path)


def convert_excel_to_csv(excel_path: str, csv_path: str) -> bool:
    """Convert Excel to CSV"""
    service = ExcelService()
    try:
        data = service.read_file(excel_path)
        return service.export_data(data, csv_path, "csv")
    except Exception:
        return False
        return service.export_data(data, csv_path, "csv")
    except Exception:
        return False
